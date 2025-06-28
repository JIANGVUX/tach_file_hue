import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import numpy as np
import math

st.set_page_config(page_title="Tách file chấm công", layout="wide")
st.title("Tách file chấm công - prodouction by Jiangvux")

def safe_excel_value(val):
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, (float, np.floating)):
        return float(round(val, 2))
    if isinstance(val, (np.integer, int)):
        return int(val)
    return val

def to_hhmm(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return ""
    if isinstance(val, (float, np.floating)) and 0 <= val < 1:
        total_minutes = int(round(val * 24 * 60))
        h, m = divmod(total_minutes, 60)
        return f"{int(h):02}:{int(m):02}"
    if hasattr(val, 'hour') and hasattr(val, 'minute'):
        return f"{val.hour:02}:{val.minute:02}"
    val_str = str(val).strip()
    if ":" in val_str:
        parts = val_str.split(":")
        if len(parts) >= 2:
            return f"{int(parts[0]):02}:{int(parts[1]):02}"
    return val_str

def get_header_row_height(header, width=8):
    lines = []
    for cell in header:
        text = str(cell.value) if cell.value else ""
        line_len = max(1, int((width - 1) * 1.5))
        num_lines = math.ceil(len(text) / line_len)
        lines.append(num_lines)
    max_lines = max(lines)
    return max(24, max_lines * 15)

def weekday_vn(date_val):
    try:
        d = pd.to_datetime(date_val, dayfirst=True)
        weekday_map = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'Chủ nhật'}
        return weekday_map[d.weekday()]
    except: return ""

uploaded_file = st.file_uploader("Chọn file Excel chấm công gốc (.xlsx)", type=["xlsx"])
if uploaded_file is not None:
    # ---- CHỈ ĐỌC DATA TỪ NHỮNG SHEET CÓ TÊN BẮT ĐẦU BẰNG "CT" ----
    xls = pd.ExcelFile(uploaded_file)
    ct_sheets = [sheet for sheet in xls.sheet_names if sheet.upper().startswith("CT")]
    if not ct_sheets:
        st.error("Không tìm thấy sheet nào bắt đầu bằng 'CT' trong file Excel!")
        st.stop()
    # Đọc & gộp lại thành 1 DataFrame
    df = pd.concat([pd.read_excel(xls, sheet_name=sheet, header=5) for sheet in ct_sheets], ignore_index=True)
    
    # ----- TỪ ĐÂY TRỞ XUỐNG: 100% LOGIC XỬ LÝ Y NHƯ CODE GỐC -----
    st.write("Tên các cột:", list(df.columns))

    # Thêm cột Thứ (trước cột Ngày)
    if "Ngày" in df.columns and "Thứ" not in df.columns:
        idx_ngay = df.columns.get_loc("Ngày")
        df.insert(idx_ngay, "Thứ", df["Ngày"].apply(weekday_vn))
    
    # Chuẩn hóa các cột thời gian về hh:mm
    cols_time = [col for col in df.columns if any(key in str(col) for key in ["Vào", "Ra"])]
    for col in cols_time:
        df[col] = df[col].apply(to_hhmm)

    # Làm tròn tự động cho các cột kiểu số
    float_cols = [col for col in df.select_dtypes(include=[np.number]).columns]
    for col in float_cols:
        df[col] = df[col].apply(lambda x: round(x, 2) if pd.notna(x) else x)

    vao_lan_1_col = next((col for col in df.columns if "Vào lần 1" in str(col)), None)
    ra_lan_2_col = next((col for col in df.columns if "Ra lần 2" in str(col)), None)
    if vao_lan_1_col is None or ra_lan_2_col is None:
        st.error('Không tìm thấy cột "Vào lần 1" hoặc "Ra lần 2" trong file!')
        st.stop()
    idx_vao_lan_1 = df.columns.get_loc(vao_lan_1_col)
    idx_ra_lan_2 = df.columns.get_loc(ra_lan_2_col)

    if "Mã NV" not in df.columns or "Họ tên" not in df.columns:
        st.error('Thiếu cột "Mã NV" hoặc "Họ tên" trong file!')
        st.stop()

    groupby_obj = list(df.groupby(['Mã NV', 'Họ tên']))
    total_nv = len(groupby_obj)
    count_nv = 0
    count_tab_vangnhat = 0
    status = st.empty()
    progress = st.progress(0)
    pale_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    output = BytesIO()
    wb_new = openpyxl.Workbook()
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    for (ma_nv, ho_ten), group in groupby_obj:
        region = group.iloc[:, idx_vao_lan_1:]
        arr = pd.Series(region.values.ravel()).astype(str).str.strip()
        arr = arr[~arr.isin(["", "nan", "NaT", "None"])]
        if arr.empty:
            continue

        count_nv += 1
        status.info(f"Đang xử lý nhân viên thứ {count_nv}/{total_nv}: **{ma_nv} - {ho_ten}**")
        progress.progress(count_nv / total_nv)

        sheet_name = f"{ma_nv}_{ho_ten}".replace(" ", "_").replace("/", "_")[:31]
        ws_nv = wb_new.create_sheet(title=sheet_name)
        ws_nv.append([str(col) for col in group.columns])
        for row in group.itertuples(index=False):
            ws_nv.append([safe_excel_value(cell) for cell in row])

        # Thêm dòng tổng cuối sheet
        total_row = []
        for col in group.columns:
            if group[col].dtype in [np.float64, np.int64, float, int]:
                tong = group[col].sum()
                if group[col].notna().sum() > 0:
                    total_row.append(round(tong, 2))
                else:
                    total_row.append("")
            else:
                total_row.append("TỔNG" if col == group.columns[0] else "")
        ws_nv.append(total_row)
        last_row = ws_nv.max_row
        for cell in ws_nv[last_row]:
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Định dạng header
        header_row = ws_nv[1]
        header_fill = PatternFill(start_color="FF8C1A", end_color="FF8C1A", fill_type="solid")
        for cell in header_row:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        ws_nv.freeze_panes = "A2"

        for i, column_cells in enumerate(ws_nv.columns):
            if i >= idx_vao_lan_1:
                ws_nv.column_dimensions[column_cells[0].column_letter].width = 8
            else:
                length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                ws_nv.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 35)
        ws_nv.row_dimensions[1].height = get_header_row_height(header_row, width=8)

        # Đánh dấu tab vàng nhạt nếu có ô bôi vàng nhạt bên trong
        co_boi_vangnhat = False

        for idx, row in enumerate(ws_nv.iter_rows(min_row=2, max_row=ws_nv.max_row-1), start=0):
            row_data = group.iloc[idx]
            region_row = row_data.iloc[idx_vao_lan_1:idx_ra_lan_2 + 1]
            values_in_region = [v for v in region_row if not (pd.isna(v) or str(v).strip() in ["", "nan", "NaT", "None"])]
            if len(values_in_region) > 1:
                for offset, value in enumerate(region_row):
                    cell = ws_nv.cell(row=idx + 2, column=idx_vao_lan_1 + 1 + offset)
                    if pd.isna(value) or str(value).strip() in ["", "nan", "NaT", "None"]:
                        cell.fill = pale_yellow_fill
                        co_boi_vangnhat = True
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        if co_boi_vangnhat:
            ws_nv.sheet_properties.tabColor = "FFFACD"
            count_tab_vangnhat += 1

    wb_new.save(output)
    output.seek(0)
    status.success(f"✅ Đã xử lý xong {count_nv} nhân viên hợp lệ!")
    progress.empty()
    st.success(f"Đã tách xong! Tổng số nhân viên được tách sheet: **{count_nv}**")
    st.info(f"Số nhân viên có sheet bị bôi vàng nhạt (tab vàng nhạt): **{count_tab_vangnhat}**")
    st.download_button(
        "Tải file Excel tổng hợp rồi cảm ơn anh Jiang đi",
        output,
        "output_tong_hop.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Đưa file lên để tách nhé!")
