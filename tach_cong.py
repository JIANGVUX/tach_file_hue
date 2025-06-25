import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import os

# Đọc file gốc, header dòng 6 (index 5)
file_path = '1.xlsx'
df = pd.read_excel(file_path, sheet_name=0, header=5)

# Xác định cột 'Vào lần 1'
vao_lan_1_col = None
for col in df.columns:
    if "Vào lần 1" in str(col):
        vao_lan_1_col = col
        break
if vao_lan_1_col is None:
    raise Exception('Không tìm thấy cột "Vào lần 1" trong file!')

# Hàm kiểm tra từ "Vào lần 1" trở đi có dữ liệu
def co_du_lieu_tu_vao_lan_1(row):
    idx = df.columns.get_loc(vao_lan_1_col)
    return any([not pd.isna(cell) and str(cell).strip() != '' for cell in row[idx:]])

# Lọc dòng hợp lệ
df_filtered = df[df.apply(co_du_lieu_tu_vao_lan_1, axis=1)]
df_filtered = df_filtered[df_filtered['Mã NV'].notna() & df_filtered['Họ tên'].notna()]

# Tạo thư mục xuất kết quả nếu muốn (tùy chọn)
output_folder = 'output'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Hàm tự động format file Excel
def auto_format_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    for ws in wb.worksheets:
        # Wrap text và căn giữa tiêu đề
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.font = Font(bold=True)
        # Wrap text cho toàn bộ sheet
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        # Auto width
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(file_path)

# Xuất file từng người
for (ma_nv, ho_ten), group in df_filtered.groupby(['Mã NV', 'Họ tên']):
    if len(group) == 0:
        continue
    file_name = f'{ma_nv}_{ho_ten}'.replace(" ", "_").replace("/", "_")
    path = os.path.join(output_folder, f'{file_name}.xlsx')
    group.to_excel(path, index=False)
    auto_format_excel(path)
    print(f'Đã xuất: {path}')

print("Hoàn tất!")
